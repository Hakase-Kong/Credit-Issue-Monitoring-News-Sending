import nltk

try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

try:
    nltk.data.find('tokenizers/punkt_tab')
except LookupError:
    nltk.download('punkt_tab')

import streamlit as st
import pandas as pd
from io import BytesIO
import requests
import re
import os
from datetime import datetime
import telepot
from openai import OpenAI
import newspaper  # newspaper4k
import openpyxl
from openpyxl import load_workbook

# --- CSS: 체크박스와 기사 사이 gap 최소화 ---
st.markdown("""
<style>
[data-testid="column"] > div {
    gap: 0rem !important;
}
.stMultiSelect [data-baseweb="tag"] {
    background-color: #ff5c5c !important;
    color: white !important;
    border: none !important;
    font-weight: bold;
}
.sentiment-badge {
    display: inline-block;
    padding: 0.08em 0.6em;
    margin-left: 0.2em;
    border-radius: 0.8em;
    font-size: 0.85em;
    font-weight: bold;
    vertical-align: middle;
}
.sentiment-positive { background: #2ecc40; color: #fff; }
.sentiment-neutral { background: #0074d9; color: #fff; }
.sentiment-negative { background: #ff4136; color: #fff; }
</style>
""", unsafe_allow_html=True)

# 세션 상태 변수 초기화
if "favorite_keywords" not in st.session_state:
    st.session_state.favorite_keywords = set()
if "search_results" not in st.session_state:
    st.session_state.search_results = {}
if "show_limit" not in st.session_state:
    st.session_state.show_limit = {}
if "search_triggered" not in st.session_state:
    st.session_state.search_triggered = False

# 대분류/소분류 카테고리
favorite_categories = {
    "국/공채": [],
    "공공기관": [],
    "보험사": ["현대해상", "농협생명", "메리츠화재", "교보생명", "삼성화재", "삼성생명", "신한라이프", "흥국생명", "동양생명", "미래에셋생명"],
    "5대금융지주": ["신한금융", "하나금융", "KB금융", "농협금융", "우리금융"],
    "5대시중은행": ["농협은행", "국민은행", "신한은행", "우리은행", "하나은행"],
    "카드사": ["KB국민카드", "현대카드", "신한카드", "비씨카드", "삼성카드"],
    "캐피탈": ["한국캐피탈", "현대캐피탈"],
    "지주사": ["SK이노베이션", "GS에너지", "SK", "GS"],
    "에너지": ["SK가스", "GS칼텍스", "S-Oil", "SK에너지", "SK앤무브", "코리아에너지터미널"],
    "발전": ["GS파워", "GSEPS", "삼천리"],
    "자동차": ["LG에너지솔루션", "한온시스템", "포스코퓨처엠", "한국타이어"],
    "전기/전자": ["SK하이닉스", "LG이노텍", "LG전자", "LS일렉트릭"],
    "소비재": ["이마트", "LF", "CJ제일제당", "SK네트웍스", "CJ대한통운"],
    "비철/철강": ["포스코", "현대제철", "고려아연"],
    "석유화학": ["LG화학", "SK지오센트릭"],
    "건설": ["포스코이앤씨"],
    "특수채": ["주택도시보증공사", "기업은행"]
}
major_categories = list(favorite_categories.keys())
sub_categories = {cat: favorite_categories[cat] for cat in major_categories}
all_fav_keywords = sorted(set(
    kw for cat in favorite_categories.values() for kw in cat if kw not in ["테스트1", "테스트2", "테스트3"]
))

st.set_page_config(layout="wide")
st.markdown("<h1 style='color:#1a1a1a; margin-bottom:0.5rem;'>📊 Credit Issue Monitoring</h1>", unsafe_allow_html=True)

# -- 검색창/검색 버튼 한 줄 배치
search_col, button_col = st.columns([7, 1])
with search_col:
    keywords_input = st.text_input("키워드 (예: 삼성, 한화)", value="", key="keyword_input")
with button_col:
    search_clicked = st.button("검색", use_container_width=True)

# -- 즐겨찾기 카테고리 선택/검색 버튼 한 줄 배치
st.markdown("**⭐ 즐겨찾기 카테고리 선택**")
cat_col, btn_col = st.columns([5, 1])
with cat_col:
    selected_categories = st.multiselect("카테고리 선택 시 자동으로 즐겨찾기 키워드에 반영됩니다.", major_categories)
    for cat in selected_categories:
        st.session_state.favorite_keywords.update(favorite_categories[cat])
with btn_col:
    category_search_clicked = st.button("🔍 검색", use_container_width=True)

# -- 즐겨찾기에서 검색/버튼 한 줄 배치
fav_col, fav_btn_col = st.columns([5, 1])
with fav_col:
    fav_selected = st.multiselect("⭐ 즐겨찾기에서 검색", all_fav_keywords, default=[])
with fav_btn_col:
    fav_search_clicked = st.button("⭐ 즐겨찾기로 검색", use_container_width=True)

# 날짜 입력
date_col1, date_col2 = st.columns([1, 1])
with date_col1:
    start_date = st.date_input("시작일")
with date_col2:
    end_date = st.date_input("종료일")

# 신용위험 필터 옵션
with st.expander("🛡️ 신용위험 필터 옵션", expanded=True):
    use_credit_filter = st.checkbox("이 필터 적용", value=False, key="use_credit_filter")
    credit_keywords = [
        "신용등급", "신용평가", "하향", "상향", "강등", "조정", "부도",
        "파산", "디폴트", "채무불이행", "적자", "영업손실", "현금흐름", "자금난",
        "재무위험", "부정적 전망", "긍정적 전망", "기업회생", "워크아웃", "구조조정", "자본잠식"
    ]
    credit_filter_keywords = st.multiselect(
        "신용위험 관련 키워드 (하나 이상 선택)",
        options=credit_keywords,
        default=credit_keywords,
        key="credit_filter"
    )

# 키워드 필터 옵션 (기본 해제)
with st.expander("🔍 키워드 필터 옵션", expanded=True):
    require_keyword_in_title = st.checkbox("기사 제목에 키워드가 포함된 경우만 보기", value=False)

# 산업별 필터 옵션 (박스형태, 한 줄에 배치, 태그 UI, 모두 선택, 체크박스)
with st.expander("🏭 산업별 필터 옵션", expanded=True):
    use_industry_filter = st.checkbox("이 필터 적용", value=False, key="use_industry_filter")
    col_major, col_sub = st.columns([1, 2])
    with col_major:
        selected_major = st.selectbox("대분류(산업)", major_categories, key="industry_major")
    with col_sub:
        selected_sub = st.multiselect(
            "소분류(필터 키워드)",
            sub_categories[selected_major],
            default=sub_categories[selected_major],
            key="industry_sub"
        )

# 재무위험 필터 옵션 (모두 선택, 체크박스)
with st.expander("💰 재무위험 필터 옵션", expanded=True):
    use_finance_filter = st.checkbox("이 필터 적용", value=False, key="use_finance_filter")
    finance_keywords = ["자산", "총자산", "부채", "자본", "매출", "비용", "영업이익", "순이익"]
    finance_filter_keywords = st.multiselect(
        "재무위험 관련 키워드",
        options=finance_keywords,
        default=finance_keywords,
        key="finance_filter"
    )

# 법/정책 위험 필터 옵션 (모두 선택, 체크박스)
with st.expander("⚖️ 법/정책 위험 필터 옵션", expanded=True):
    use_law_filter = st.checkbox("이 필터 적용", value=False, key="use_law_filter")
    law_keywords = ["테스트1", "테스트2", "테스트3"]
    law_filter_keywords = st.multiselect(
        "법/정책 위험 관련 키워드",
        options=law_keywords,
        default=law_keywords,
        key="law_filter"
    )

# --- 본문 추출 함수(요청대로 단순화) ---
def extract_article_text(url):
    try:
        article = newspaper.article(url)
        article.download()
        article.parse()
        return article.text
    except Exception as e:
        return f"본문 추출 오류: {e}"

# --- OpenAI 요약/감성분석 함수 ---
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY)

def detect_lang(text):
    return "ko" if re.search(r"[가-힣]", text) else "en"

def summarize_and_sentiment_with_openai(text):
    if not OPENAI_API_KEY:
        return "OpenAI API 키가 설정되지 않았습니다.", None, None, None
    lang = detect_lang(text)
    if lang == "ko":
        prompt = (
            "아래 기사 본문을 요약하고 감성분석을 해줘.\n\n"
            "- [한 줄 요약]: 기사 전체 내용을 한 문장으로 요약\n"
            "- [요약본]: 기사 내용을 2~3 문단(각 문단 2~4문장)으로, 핵심 내용을 충분히 파악할 수 있게 요약\n"
            "- [감성]: 기사 전체의 감정을 긍정/부정/중립 중 하나로만 답해줘. "
            "만약 파산, 자금난, 회생, 적자, 구조조정, 영업손실, 부도, 채무불이행, 경영 위기 등 부정적 사건이 중심이면 반드시 '부정'으로 답해줘.\n"
            "광고, 배너, 추천기사, 서비스 안내 등 기사 본문과 무관한 내용은 모두 요약과 감성분석에서 제외.\n\n"
            "아래 포맷으로 답변해줘:\n"
            "[한 줄 요약]: (여기에 한 줄 요약)\n"
            "[요약본]: (여기에 여러 문단 요약)\n"
            "[감성]: (긍정/부정/중립 중 하나만)\n\n"
            "[기사 본문]\n" + text
        )
    else:
        prompt = (
            "Summarize the following news article and analyze its sentiment.\n\n"
            "- [One-line Summary]: Summarize the entire article in one sentence.\n"
            "- [Summary]: Summarize the article in 2–3 paragraphs (each 2–4 sentences), so that the main content is well understood.\n"
            "- [Sentiment]: Classify the overall sentiment as one of: positive, negative, or neutral. "
            "If the article centers on bankruptcy, financial distress, restructuring, insolvency, operating loss, default, or management crisis, you must answer 'negative'.\n"
            "Exclude any advertisements, banners, recommended articles, or unrelated content.\n\n"
            "Respond in this format:\n"
            "[One-line Summary]: (your one-line summary)\n"
            "[Summary]: (your multi-paragraph summary)\n"
            "[Sentiment]: (positive/negative/neutral only)\n\n"
            "[ARTICLE]\n" + text
        )
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": prompt}
        ],
        max_tokens=1024,
        temperature=0.3
    )
    answer = response.choices[0].message.content.strip()
    if lang == "ko":
        m1 = re.search(r"\[한 줄 요약\]:\s*(.+)", answer)
        m2 = re.search(r"\[요약본\]:\s*([\s\S]+?)(?:\[감성\]:|$)", answer)
        m3 = re.search(r"\[감성\]:\s*(.+)", answer)
    else:
        m1 = re.search(r"\[One-line Summary\]:\s*(.+)", answer)
        m2 = re.search(r"\[Summary\]:\s*([\s\S]+?)(?:\[Sentiment\]:|$)", answer)
        m3 = re.search(r"\[Sentiment\]:\s*(.+)", answer)
    one_line = m1.group(1).strip() if m1 else ""
    summary = m2.group(1).strip() if m2 else answer
    sentiment = m3.group(1).strip() if m3 else ""
    return one_line, summary, sentiment, text

# --- 텔레그램 클래스 ---
NAVER_CLIENT_ID = "_qXuzaBGk_jQesRRPRvu"
NAVER_CLIENT_SECRET = "lZc2gScgNq"
TELEGRAM_TOKEN = "7033950842:AAFk4pSb5qtNj435Gf2B5-rPlFrlNqhZFuQ"
TELEGRAM_CHAT_ID = "-1002404027768"

class Telegram:
    def __init__(self):
        self.bot = telepot.Bot(TELEGRAM_TOKEN)
        self.chat_id = TELEGRAM_CHAT_ID

    def send_message(self, message):
        self.bot.sendMessage(self.chat_id, message, parse_mode="Markdown", disable_web_page_preview=True)

# --- 뉴스 API 함수 (네이버/GNews) ---
def filter_by_issues(title, desc, selected_keywords, enable_credit_filter, credit_filter_keywords, require_keyword_in_title=False):
    if require_keyword_in_title and selected_keywords:
        if not any(kw.lower() in title.lower() for kw in selected_keywords):
            return False
    if enable_credit_filter and not is_credit_risk_news(title + " " + desc, credit_filter_keywords):
        return False
    return True

def is_credit_risk_news(text, keywords):
    return any(kw in text for kw in keywords)

def fetch_naver_news(query, start_date=None, end_date=None, enable_credit_filter=True, credit_filter_keywords=None, limit=100, require_keyword_in_title=False):
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET
    }
    articles = []
    for page in range(1, 6):
        if len(articles) >= limit:
            break
        params = {
            "query": query,
            "display": 10,
            "start": (page - 1) * 10 + 1,
            "sort": "date"
        }
        response = requests.get("https://openapi.naver.com/v1/search/news.json", headers=headers, params=params)
        if response.status_code != 200:
            break
        items = response.json().get("items", [])
        for item in items:
            title, desc = item["title"], item["description"]
            pub_date = datetime.strptime(item["pubDate"], "%a, %d %b %Y %H:%M:%S %z").date()
            if start_date and pub_date < start_date:
                continue
            if end_date and pub_date > end_date:
                continue
            if not filter_by_issues(title, desc, [query], use_credit_filter, credit_filter_keywords, require_keyword_in_title):
                continue
            articles.append({
                "title": re.sub("<.*?>", "", title),
                "link": item["link"],
                "date": pub_date.strftime("%Y-%m-%d"),
                "source": "Naver"
            })
    return articles[:limit]

def fetch_gnews_news(query, enable_credit_filter=True, credit_filter_keywords=None, limit=100, require_keyword_in_title=False):
    GNEWS_API_KEY = "b8c6d82bbdee9b61d2b9605f44ca8540"
    articles = []
    try:
        url = f"https://gnews.io/api/v4/search"
        params = {
            "q": query,
            "lang": "en",
            "token": GNEWS_API_KEY,
            "max": limit
        }
        response = requests.get(url, params=params)
        if response.status_code != 200:
            st.warning(f"❌ GNews 요청 실패 - 상태 코드: {response.status_code}")
            return []
        data = response.json()
        for item in data.get("articles", []):
            title = item.get("title", "")
            desc = item.get("description", "")
            if not filter_by_issues(title, desc, [query], use_credit_filter, credit_filter_keywords, require_keyword_in_title):
                continue
            pub_date = datetime.strptime(item["publishedAt"][:10], "%Y-%m-%d").date()
            articles.append({
                "title": title,
                "link": item.get("url", ""),
                "date": pub_date.strftime("%Y-%m-%d"),
                "source": "GNews"
            })
    except Exception as e:
        st.warning(f"⚠️ GNews 접근 오류: {e}")
    return articles

def is_english(text):
    return all(ord(c) < 128 for c in text if c.isalpha())

def process_keywords(keyword_list, start_date, end_date, enable_credit_filter, credit_filter_keywords, require_keyword_in_title=False):
    for k in keyword_list:
        if is_english(k):
            articles = fetch_gnews_news(k, enable_credit_filter, credit_filter_keywords, require_keyword_in_title=require_keyword_in_title)
        else:
            articles = fetch_naver_news(k, start_date, end_date, enable_credit_filter, credit_filter_keywords, require_keyword_in_title=require_keyword_in_title)
        st.session_state.search_results[k] = articles
        st.session_state.show_limit[k] = 5

def detect_lang_from_title(title):
    return "ko" if re.search(r"[가-힣]", title) else "en"

def summarize_article_from_url(article_url, title):
    try:
        full_text = extract_article_text(article_url)
        if full_text.startswith("본문 추출 오류"):
            return full_text, None, None, None
        one_line, summary, sentiment, _ = summarize_and_sentiment_with_openai(full_text)
        return one_line, summary, sentiment, full_text
    except Exception as e:
        return f"요약 오류: {e}", None, None, None

# OR 조건 필터링 함수
def or_keyword_filter(article, *keyword_lists):
    text = (article.get("title", "") + " " + article.get("description", "") + " " + article.get("full_text", ""))
    for keywords in keyword_lists:
        if any(kw in text for kw in keywords if kw):
            return True
    return False

# 실제 뉴스 검색/필터링/요약/감성분석 실행
search_clicked = False
if keywords_input:
    keyword_list = [k.strip() for k in keywords_input.split(",") if k.strip()]
    if len(keyword_list) > 10:
        st.warning("키워드는 최대 10개까지 입력 가능합니다.")
    else:
        search_clicked = True

if search_clicked or st.session_state.get("search_triggered"):
    keyword_list = [k.strip() for k in keywords_input.split(",") if k.strip()]
    if len(keyword_list) > 10:
        st.warning("키워드는 최대 10개까지 입력 가능합니다.")
    else:
        with st.spinner("뉴스 검색 중..."):
            process_keywords(keyword_list, start_date, end_date, use_credit_filter, credit_filter_keywords, require_keyword_in_title)
    st.session_state.search_triggered = False

if fav_search_clicked and fav_selected:
    with st.spinner("뉴스 검색 중..."):
        process_keywords(fav_selected, start_date, end_date, use_credit_filter, credit_filter_keywords, require_keyword_in_title)

if category_search_clicked and selected_categories:
    with st.spinner("뉴스 검색 중..."):
        keywords = set()
        for cat in selected_categories:
            keywords.update(favorite_categories[cat])
        process_keywords(
            sorted(keywords),
            start_date,
            end_date,
            use_credit_filter,
            credit_filter_keywords,
            require_keyword_in_title
        )

# 필터링: OR 조건(필터별 체크박스에 따라 적용)
def article_passes_all_filters(article):
    filters = []
    if use_credit_filter:
        filters.append(credit_filter_keywords)
    if use_industry_filter:
        filters.append(selected_sub)
    if use_finance_filter:
        filters.append(finance_filter_keywords)
    if use_law_filter:
        filters.append(law_filter_keywords)
    if filters:
        return or_keyword_filter(article, *filters)
    else:
        return True

# --- 엑셀 업데이트 함수 (openpyxl) ---
def update_excel(selected_data, template_path):
    wb = load_workbook(template_path)
    ws = wb.active
    # 회사명 → 행번호 매핑 (엑셀의 3행부터 데이터 시작, 회사명은 D열(5번째))
    company_col = 4  # D열(0부터 시작)
    company_to_row = {}
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=company_col).value
        if name:
            company_to_row[name.replace(" ", "")] = row
    # J(10), L(12)열에 하이퍼링크 업데이트
    for item in selected_data:
        name = item["회사명"].replace(" ", "")
        if name in company_to_row and item["요약"] and item["링크"]:
            if item["sentiment"] == "긍정":
                cell = ws.cell(row=company_to_row[name], column=10)  # J열
            elif item["sentiment"] == "부정":
                cell = ws.cell(row=company_to_row[name], column=12)  # L열
            else:
                continue
            cell.value = item["요약"]
            cell.hyperlink = item["링크"]
            cell.style = "Hyperlink"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 요약/감성분석/기사선택/엑셀 저장 UI ---
def render_articles_with_single_summary_and_telegram(results, show_limit):
    SENTIMENT_CLASS = {
        "긍정": "sentiment-positive",
        "부정": "sentiment-negative",
        "중립": "sentiment-neutral"
    }
    summary_data = []
    checked_list = []

    # 선택 상태를 세션에 저장 (체크박스 상태 유지)
    if "article_checked" not in st.session_state:
        st.session_state.article_checked = {}

    # 2단 컬럼 레이아웃 (왼쪽: 기사리스트, 오른쪽: 선택된 기사 요약/감성)
    col_list, col_summary = st.columns([1, 1])

    with col_list:
        st.markdown("### 기사 요약 결과 (엑셀 저장할 기사 선택)")
        for keyword, articles in results.items():
            for idx, article in enumerate(articles[:show_limit.get(keyword, 5)]):
                key = f"{keyword}_{idx}"
                cache_key = f"summary_{key}"
                # 감성분석 캐싱 (기사 리스트에 바로 보여주기 위해)
                if cache_key not in st.session_state:
                    one_line, summary, sentiment, full_text = summarize_article_from_url(article['link'], article['title'])
                    st.session_state[cache_key] = (one_line, summary, sentiment, full_text)
                else:
                    one_line, summary, sentiment, full_text = st.session_state[cache_key]
                sentiment_label = sentiment if sentiment else "분석중"
                sentiment_class = SENTIMENT_CLASS.get(sentiment_label, "sentiment-neutral")
                # 기사 제목 옆에 감성 결과를 괄호+색상 뱃지로 바로 표시
                md_line = (
                    f"[{keyword}] "
                    f"[{article['title']}]({article['link']}) "
                    f"<span class='sentiment-badge {sentiment_class}'>({sentiment_label})</span> "
                    f"({article['date']} | {article['source']})"
                )
                cols = st.columns([0.04, 0.96])
                with cols[0]:
                    checked = st.checkbox("", value=st.session_state.article_checked.get(key, False), key=f"news_{key}")
                with cols[1]:
                    st.markdown(md_line, unsafe_allow_html=True)
                st.session_state.article_checked[key] = checked

    with col_summary:
        st.markdown("### 선택된 기사 요약/감성분석")
        selected_articles = []
        for keyword, articles in results.items():
            for idx, article in enumerate(articles[:show_limit.get(keyword, 5)]):
                key = f"{keyword}_{idx}"
                cache_key = f"summary_{key}"
                if st.session_state.article_checked.get(key, False):
                    if cache_key not in st.session_state:
                        one_line, summary, sentiment, full_text = summarize_article_from_url(article['link'], article['title'])
                        st.session_state[cache_key] = (one_line, summary, sentiment, full_text)
                    else:
                        one_line, summary, sentiment, full_text = st.session_state[cache_key]
                    selected_articles.append({
                        "회사명": keyword,
                        "기사제목": article['title'],
                        "요약": one_line,
                        "full_summary": summary,
                        "sentiment": sentiment,
                        "링크": article['link'],
                        "date": article['date'],
                        "source": article['source']
                    })
                    # 요약/감성분석 결과 출력
                    st.markdown(f"#### [{article['title']}]({article['link']}) <span class='sentiment-badge {SENTIMENT_CLASS.get(sentiment, 'sentiment-neutral')}'>({sentiment})</span>", unsafe_allow_html=True)
                    st.markdown(f"- **날짜/출처:** {article['date']} | {article['source']}")
                    st.markdown(f"- **한 줄 요약:** {one_line}")
                    st.markdown(f"- **요약본:** {summary}")
                    st.markdown("---")
        summary_data = selected_articles

        st.write(f"선택된 기사 개수: {len(summary_data)}")

        # 3. 엑셀 템플릿 업로드 및 저장
        st.markdown("#### 기존 엑셀 템플릿 업로드")
        uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요(기존 템플릿)", type=["xlsx"])
        if uploaded_file is not None and summary_data:
            if st.button("선택 기사 엑셀로 저장"):
                excel_bytes = update_excel(summary_data, uploaded_file)
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=excel_bytes.getvalue(),
                    file_name="뉴스요약_업데이트.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        elif uploaded_file is not None:
            st.info("엑셀로 저장할 기사를 먼저 선택하세요.")

if st.session_state.search_results:
    filtered_results = {}
    for keyword, articles in st.session_state.search_results.items():
        filtered_articles = [a for a in articles if article_passes_all_filters(a)]
        if filtered_articles:
            filtered_results[keyword] = filtered_articles
    render_articles_with_single_summary_and_telegram(filtered_results, st.session_state.show_limit)
